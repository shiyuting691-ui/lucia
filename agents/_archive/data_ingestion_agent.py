"""
DataIngestionAgent — CSV/Excel/JSON 数据导入
支持导入：订单(orders)、咨询(leads)、学校节点(school_calendar)
后续可扩展至 CRM / 飞书表格 API
"""
import logging
from pathlib import Path
from datetime import datetime
from database import save_order, save_lead, save_school_calendar, save_yearly_pattern

logger = logging.getLogger(__name__)


# ── 字段别名映射（兼容常见中英文表头）──────────────────────────
ORDER_ALIASES = {
    "order_date":   ["order_date", "订单日期", "日期", "date"],
    "customer_id":  ["customer_id", "客户ID", "客户编号", "id"],
    "school":       ["school", "学校", "学校名称"],
    "country":      ["country", "国家", "地区"],
    "major":        ["major", "专业"],
    "course_code":  ["course_code", "课程代码", "课程", "course"],
    "product":      ["product", "产品", "产品ID", "product_id"],
    "service_type": ["service_type", "服务类型", "辅导类型"],
    "deadline":     ["deadline", "截止日期", "考试日期", "ddl"],
    "amount":       ["amount", "金额", "成交金额", "price"],
    "sales_owner":  ["sales_owner", "销售", "顾问", "负责人"],
    "status":       ["status", "状态", "订单状态"],
}

LEAD_ALIASES = {
    "inquiry_date":     ["inquiry_date", "咨询日期", "日期", "date"],
    "customer_name":    ["customer_name", "客户姓名", "姓名", "name"],
    "school":           ["school", "学校"],
    "country":          ["country", "国家"],
    "major":            ["major", "专业"],
    "course_code":      ["course_code", "课程代码", "课程"],
    "product_interest": ["product_interest", "意向产品", "产品", "product"],
    "pain_point":       ["pain_point", "痛点", "问题", "需求"],
    "deadline":         ["deadline", "截止日期", "ddl"],
    "quoted_price":     ["quoted_price", "报价", "price"],
    "deal_status":      ["deal_status", "成交状态", "跟进状态", "status"],
    "lost_reason":      ["lost_reason", "流失原因", "未成交原因"],
    "sales_owner":      ["sales_owner", "销售", "顾问", "负责人"],
    "source_channel":   ["source_channel", "来源渠道", "渠道", "channel"],
}

CALENDAR_ALIASES = {
    "school":        ["school", "学校"],
    "country":       ["country", "国家"],
    "academic_year": ["academic_year", "学年"],
    "term":          ["term", "学期"],
    "event_type":    ["event_type", "节点类型", "类型"],
    "event_name":    ["event_name", "节点名称", "事件"],
    "start_date":    ["start_date", "开始日期", "start"],
    "end_date":      ["end_date", "结束日期", "end"],
    "confidence":    ["confidence", "置信度"],
    "source":        ["source", "来源"],
    "notes":         ["notes", "备注"],
}


def _normalize_row(row: dict, aliases: dict) -> dict:
    """将任意列名的行映射到标准字段名"""
    result = {}
    row_lower = {k.lower().strip(): v for k, v in row.items()}
    for field, candidates in aliases.items():
        for cand in candidates:
            val = row_lower.get(cand.lower())
            if val is not None and str(val).strip() not in ("", "nan", "None", "NaN"):
                result[field] = str(val).strip()
                break
    return result


def _load_file(path: str) -> list[dict]:
    """加载 CSV / Excel / JSON 文件，返回 list[dict]"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"文件不存在：{path}")

    suffix = p.suffix.lower()
    if suffix == ".json":
        import json
        data = json.loads(p.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else [data]

    elif suffix == ".csv":
        import csv
        rows = []
        with open(p, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
        return rows

    elif suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("导入 Excel 需要安装 openpyxl：uv add openpyxl")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        return rows

    else:
        raise ValueError(f"不支持的文件格式：{suffix}（支持 .csv / .xlsx / .json）")


class DataIngestionAgent:
    def __init__(self, config: dict = None):
        self.config = config or {}

    def ingest_orders(self, file_path: str) -> dict:
        """导入订单 CSV/Excel"""
        try:
            rows = _load_file(file_path)
            saved = skipped = 0
            for row in rows:
                norm = _normalize_row(row, ORDER_ALIASES)
                if not norm.get("order_date") and not norm.get("school"):
                    skipped += 1
                    continue
                norm["source_file"] = file_path
                try:
                    save_order(norm)
                    saved += 1
                except Exception as e:
                    logger.warning(f"订单行跳过：{e} | {norm}")
                    skipped += 1
            msg = f"订单导入完成：成功 {saved} 条，跳过 {skipped} 条"
            logger.info(msg)
            return {"saved": saved, "skipped": skipped, "message": msg}
        except Exception as e:
            logger.error(f"ingest_orders error: {e}")
            return {"saved": 0, "skipped": 0, "error": str(e)}

    def ingest_leads(self, file_path: str) -> dict:
        """导入咨询/线索 CSV/Excel"""
        try:
            rows = _load_file(file_path)
            saved = skipped = 0
            for row in rows:
                norm = _normalize_row(row, LEAD_ALIASES)
                if not norm.get("inquiry_date") and not norm.get("school"):
                    skipped += 1
                    continue
                norm["source_file"] = file_path
                try:
                    save_lead(norm)
                    saved += 1
                except Exception as e:
                    logger.warning(f"咨询行跳过：{e} | {norm}")
                    skipped += 1
            msg = f"咨询导入完成：成功 {saved} 条，跳过 {skipped} 条"
            logger.info(msg)
            return {"saved": saved, "skipped": skipped, "message": msg}
        except Exception as e:
            logger.error(f"ingest_leads error: {e}")
            return {"saved": 0, "skipped": 0, "error": str(e)}

    def ingest_school_calendar(self, file_path: str) -> dict:
        """导入学校节点 CSV/Excel"""
        try:
            rows = _load_file(file_path)
            saved = skipped = 0
            for row in rows:
                norm = _normalize_row(row, CALENDAR_ALIASES)
                if not norm.get("school") or not norm.get("event_type"):
                    skipped += 1
                    continue
                try:
                    save_school_calendar(norm)
                    saved += 1
                except Exception as e:
                    logger.warning(f"节点行跳过：{e}")
                    skipped += 1
            msg = f"学校节点导入完成：成功 {saved} 条，跳过 {skipped} 条"
            logger.info(msg)
            return {"saved": saved, "skipped": skipped, "message": msg}
        except Exception as e:
            logger.error(f"ingest_school_calendar error: {e}")
            return {"saved": 0, "skipped": 0, "error": str(e)}
